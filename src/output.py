import json
import logging
import datetime
import time
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from utils.config_loader import cfg

logger = logging.getLogger(__name__)

# ==================== КОНСТАНТЫ ====================
COLUMNS_ORDER = ["Задача", "В", "Ответственный", "Срок", "Комментарии", 
                 "Пометка ЕС", "Пометка ЕКс", "Пометка МН", "Пометка НС"]

# Стандартные задачи, которые всегда должны быть в отчёте
DEFAULT_TASKS = [
    {"Задача": "Прослушивание звонков админов...", "В": "", "Ответственный": "Глухова", 
     "Срок": "Еженедельно", "Комментарии": "", "Пометка ЕС": "", "Пометка ЕКс": "", 
     "Пометка МН": "", "Пометка НС": ""},
    {"Задача": "Резюме по отправленным в работу звонкам...", "В": "", "Ответственный": "Савина ЕС", 
     "Срок": "Еженедельно", "Комментарии": "", "Пометка ЕС": "", "Пометка ЕКс": "", 
     "Пометка МН": "", "Пометка НС": ""},
    {"Задача": "Контроль постановки отзвонов...", "В": "", "Ответственный": "Савина Е.С. / Глухова ЕС", 
     "Срок": "Еженедельно", "Комментарии": "", "Пометка ЕС": "", "Пометка ЕКс": "", 
     "Пометка МН": "", "Пометка НС": ""},
]

# ==================== ФОРМАТИРОВАНИЕ ====================
def format_excel_sheet(worksheet):
    """Применяет стили к листу Excel"""
    # Более полная настройка ширины колонок
    dims = {
        'A': 45,  # Задача
        'B': 5,   # В
        'C': 25,  # Ответственный
        'D': 15,  # Срок
        'E': 50,  # Комментарии
        'F': 18, 'G': 18, 'H': 18, 'I': 18  # Пометки
    }
    
    for col_letter in [chr(65 + i) for i in range(len(COLUMNS_ORDER))]:
        width = dims.get(col_letter, 12)
        worksheet.column_dimensions[col_letter].width = width
    
    # Стили заголовка
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True)
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Стили данных
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')


# ==================== ОСНОВНАЯ ФУНКЦИЯ ====================
def write_output(data, video_date=None):
    """Основная функция вывода результатов"""
    rows = data.get('rows', []) if isinstance(data, dict) else data

    # 1. Консольный лог
    logger.info("╒══ РЕЗУЛЬТАТЫ АНАЛИЗА")
    if not rows:
        logger.info("│   [ EMPTY ] Задач не обнаружено.")
    else:
        for idx, item in enumerate(rows, 1):
            assignee = item.get('assignee_code') or item.get('assignee') or '??'
            task = item.get('task') or 'Без названия'
            logger.info(f"│   {idx:>2}. [{assignee.upper()}] {task}")
    logger.info(f"┕━━ Всего задач от Gemini: {len(rows)}")

    # 2. Excel
    if str(cfg.get('OUTPUT', 'enable_excel', fallback='True')).lower() in ('true', '1', 'yes', 'on'):
        _write_to_excel(rows, video_date)

    # 3. JSON Debug
    _save_debug_json(rows)


def _write_to_excel(gemini_rows, video_date=None):
    """Запись в Excel с DEFAULT_TASKS и улучшенной обработкой"""
    excel_path = Path(cfg.get('PATHS', 'output_excel', fallback='output/Report.xlsx'))
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    
    if not video_date:
        video_date = datetime.datetime.now()
    
    sheet_name = video_date.strftime('%d.%m.%Y')
    default_deadline = (video_date + datetime.timedelta(days=7)).strftime("%d.%m.%Y")

    # Формируем все строки
    all_rows = []

    # 1. Стандартные задачи
    for task in DEFAULT_TASKS:
        all_rows.append(task.copy())

    # 2. Задачи от Gemini
    for item in gemini_rows or []:
        deadline = item.get("deadline")
        if not deadline or str(deadline).lower() in ["нет", "none", "—", "null", ""]:
            deadline = default_deadline
            
        all_rows.append({
            "Задача": item.get("task", ""),
            "В": "",
            "Ответственный": item.get("assignee_name") or item.get("assignee", ""),
            "Срок": deadline,
            "Комментарии": item.get("details") or item.get("report_text", "") or "",
            "Пометка ЕС": "", "Пометка ЕКс": "", "Пометка МН": "", "Пометка НС": ""
        })

    if not all_rows:
        logger.info("│   Нет данных для записи в Excel.")
        return

    df_new = pd.DataFrame(all_rows).reindex(columns=COLUMNS_ORDER)

    # Запись с ретраями
    max_retries = 5
    for attempt in range(max_retries):
        try:
            file_exists = excel_path.exists()
            mode = 'a' if file_exists else 'w'
            
            write_params = {'engine': 'openpyxl', 'mode': mode}
            if file_exists:
                write_params['if_sheet_exists'] = 'replace'
            
            with pd.ExcelWriter(excel_path, **write_params) as writer:
                df_new.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Перемещаем новый лист в начало (как в старой версии)
                wb = writer.book
                if len(wb.sheetnames) > 1:
                    sheet = wb[sheet_name]
                    wb._sheets.remove(sheet)
                    wb._sheets.insert(0, sheet)
                    wb.active = 0
                
                format_excel_sheet(wb[sheet_name])
            
            logger.info(f"│   ╰─ [ OK ] Excel: Лист '{sheet_name}' сохранен и перемещен наверх.")
            break
            
        except PermissionError:
            if attempt < max_retries - 1:
                logger.warning(f"│   ╰─ [ WAIT ] Закройте файл {excel_path.name}! Попытка {attempt+1}/{max_retries}...")
                time.sleep(6)
            else:
                logger.error(f"│   ╰─ [ FAIL ] Не удалось записать Excel — файл занят.")
        except Exception as e:
            logger.error(f"│   ╰─ [ FAIL ] Ошибка записи Excel: {e}")
            break


def _save_debug_json(rows):
    """Сохранение отладочной информации"""
    debug_path = Path("output/last_run_debug.json")
    try:
        debug_path.parent.mkdir(parents=True, exist_ok=True)
        with debug_path.open('w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        logger.debug(f"Дебаг JSON сохранен: {debug_path}")
    except Exception as e:
        logger.debug(f"Не удалось сохранить debug JSON: {e}")